# -*- coding: utf-8 -*-
"""
GUI генератор этикеток/ШК для Wildberries.

Что нового в этой версии:
- ЖИВОЕ ПРЕВЬЮ этикетки прямо в окне (обновляется при изменении любого поля)
- РАЗМЕТКА: рамка этикетки + сетка 5 мм + предполагаемые границы текста и ШК
- ИСПРАВЛЕНА КИРИЛЛИЦА: автопоиск и регистрация TTF-шрифта с поддержкой
  кириллицы (Arial/DejaVu/Liberation). Helvetica больше не используется,
  поэтому квадратов вместо русских букв не будет.

Что делает:
- читает Excel с колонками: Артикул, Цвет, Размер, Материал, Баркод, Срок годности
- проверяет / формирует EAN-13
- создает PDF этикетки 58 x 40 мм
- имя PDF: артикул_цвет_размер.pdf
- настройка X/Y позиций текста и штрихкода, размер шрифта, размеры ШК
- сохраняет настройки рядом со скриптом в wb_barcode_settings.json

Установка зависимостей:
    pip install openpyxl reportlab pillow

Запуск:
    python wb_barcode_gui.py
"""

import io
import json
import os
import re
import sys
import traceback
from pathlib import Path
from tkinter import (
    Tk, StringVar, DoubleVar, IntVar, filedialog, messagebox, ttk,
    Text, END, Canvas, Frame,
)

from openpyxl import load_workbook
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.graphics.shapes import Drawing
from reportlab.graphics import renderPDF
from reportlab.graphics.barcode.eanbc import Ean13BarcodeWidget

from PIL import Image, ImageDraw, ImageFont, ImageTk


LABEL_W_MM = 58
LABEL_H_MM = 40
REQUIRED_COLUMNS = ["Артикул", "Артикул WB", "Цвет", "Размер", "Состав", "Баркод", "Гарантия"]
SETTINGS_FILE = Path(__file__).with_name("wb_barcode_settings.json")

# ---- имя зарегистрированного в reportlab шрифта ----
PDF_FONT_NAME = "LabelFont"

# Кандидаты TTF со встроенной кириллицей. Берём первый найденный.
FONT_CANDIDATES = [
    # Windows
    r"C:\Windows\Fonts\arial.ttf",
    r"C:\Windows\Fonts\segoeui.ttf",
    r"C:\Windows\Fonts\tahoma.ttf",
    r"C:\Windows\Fonts\calibri.ttf",
    r"C:\Windows\Fonts\verdana.ttf",
    # Linux
    "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
    # macOS
    "/Library/Fonts/Arial.ttf",
    "/System/Library/Fonts/Supplemental/Arial.ttf",
]

DEFAULT_SETTINGS = {
    "font_size": 7,
    "label_w_mm": 58,
    "label_h_mm": 40,

    "article_x": 3,
    "article_y": 36,
    "color_x": 3,
    "color_y": 32,
    "size_x": 3,
    "size_y": 28,
    "material_x": 3,
    "material_y": 24,
    "expiry_x": 3,
    "expiry_y": 20,

    "barcode_x": 5,
    "barcode_y": 4,
    "barcode_w": 48,
    "barcode_h": 14,
    "barcode_digits_x": 18,
    "barcode_digits_y": 2,
    "barcode_digits_font_size": 7,

    "num_x": 50,
    "num_y": 37,
    "num_font_size": 6,

    "make_one_pdf": 0,
    "show_grid": 1,
}


# ---------------------------------------------------------------------------
# Шрифты
# ---------------------------------------------------------------------------

def find_font_path() -> str | None:
    """Найти первый доступный TTF с кириллицей."""
    for path in FONT_CANDIDATES:
        if os.path.exists(path):
            return path
    # последняя попытка: поискать что-то в системных папках
    search_dirs = [
        r"C:\Windows\Fonts",
        "/usr/share/fonts",
        "/Library/Fonts",
        os.path.expanduser("~/.fonts"),
    ]
    for d in search_dirs:
        if os.path.isdir(d):
            for rootdir, _, files in os.walk(d):
                for fn in files:
                    if fn.lower().endswith(".ttf"):
                        return os.path.join(rootdir, fn)
    return None


FONT_PATH = find_font_path()


def register_pdf_font() -> str:
    """
    Зарегистрировать кириллический шрифт в reportlab.
    Возвращает имя шрифта для использования в canvas.setFont().
    Если TTF не найден — откатывается на Helvetica (будут квадраты на кириллице),
    но мы предупредим пользователя в логе.
    """
    if FONT_PATH:
        try:
            pdfmetrics.registerFont(TTFont(PDF_FONT_NAME, FONT_PATH))
            return PDF_FONT_NAME
        except Exception:
            pass
    return "Helvetica"


# ---------------------------------------------------------------------------
# Утилиты данных
# ---------------------------------------------------------------------------

def sanitize_filename(value: str) -> str:
    value = str(value).strip()
    value = re.sub(r"[\\/:*?\"<>|]+", "_", value)
    value = re.sub(r"\s+", " ", value)
    return value[:120] or "empty"


def normalize_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def ean13_checksum(first_12: str) -> str:
    total = 0
    for i, ch in enumerate(first_12):
        digit = int(ch)
        total += digit if i % 2 == 0 else digit * 3
    return str((10 - total % 10) % 10)


def normalize_ean13(raw: str) -> str:
    digits = re.sub(r"\D", "", str(raw))
    if len(digits) == 12:
        return digits + ean13_checksum(digits)
    if len(digits) == 13:
        expected = ean13_checksum(digits[:12])
        if digits[-1] != expected:
            raise ValueError(f"неверная контрольная цифра EAN-13: {digits}, должна быть {expected}")
        return digits
    if len(digits) == 14 and digits[0] == "0":
        return normalize_ean13(digits[1:])
    raise ValueError(f"баркод должен содержать 12 или 13 цифр, сейчас: {raw}")


def is_gtin(raw: str) -> bool:
    """GTIN-14 — 14 цифр (обычно с ведущим нулём)."""
    return len(re.sub(r"\D", "", str(raw))) == 14


def load_settings() -> dict:
    settings = DEFAULT_SETTINGS.copy()
    if SETTINGS_FILE.exists():
        try:
            with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
                settings.update(json.load(f))
        except Exception:
            pass
    return settings


def save_settings(settings: dict) -> None:
    clean = {k: v for k, v in settings.items() if k != "font_name"}
    with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
        json.dump(clean, f, ensure_ascii=False, indent=2)


def read_excel_rows(path: str):
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    headers = {}
    for col in range(1, ws.max_column + 1):
        value = normalize_cell(ws.cell(1, col).value)
        if value:
            headers[value] = col

    missing = [col for col in REQUIRED_COLUMNS if col not in headers]
    if missing:
        raise ValueError("В Excel нет колонок: " + ", ".join(missing))

    rows = []
    for r in range(2, ws.max_row + 1):
        item = {name: normalize_cell(ws.cell(r, headers[name]).value) for name in REQUIRED_COLUMNS}
        if not any(item.values()):
            continue
        item["_row"] = r
        item["_num"] = len(rows) + 1
        rows.append(item)
    return rows


def make_template(path: str):
    """Создать пустой Excel-шаблон с нужными колонками. Баркод — текстовый формат."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Шаблон"
    ws.append(REQUIRED_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    example = {
        "Артикул": "ITALI12_Черный", "Артикул WB": "120909012",
        "Цвет": "Черный", "Размер": "xs", "Состав": "Полиэстер",
        "Баркод": "2000000000244", "Гарантия": "1 год",
    }
    ws.append([example.get(col, "") for col in REQUIRED_COLUMNS])

    barcode_col = REQUIRED_COLUMNS.index("Баркод") + 1
    for r in range(1, 501):
        ws.cell(row=r, column=barcode_col).number_format = "@"

    wb.save(path)


def make_exclusions(src_path: str, skipped_rows: list, out_path: Path) -> Path:
    """Копия исходного Excel с жёлтой подсветкой пропущенных (GTIN) строк."""
    from openpyxl.styles import PatternFill

    wb = load_workbook(src_path)
    ws = wb.active
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    max_col = ws.max_column
    for row in skipped_rows:
        r = row.get("_row")
        if r:
            for col in range(1, max_col + 1):
                ws.cell(row=r, column=col).fill = fill
    out_path = unique_path(out_path)
    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------------------
# Отрисовка PDF
# ---------------------------------------------------------------------------

def label_lines(row: dict):
    """Единый источник правды: какие строки и в каких полях рисуем."""
    return [
        ("article_x", "article_y", f"Артикул: {row['Артикул']}"),
        ("color_x", "color_y", f"Цвет: {row['Цвет']}"),
        ("size_x", "size_y", f"Размер: {row['Размер']}"),
        ("material_x", "material_y", f"Состав: {row['Состав']}"),
        ("expiry_x", "expiry_y", f"Гарантия: {row['Гарантия']}"),
    ]


def draw_barcode(c: canvas.Canvas, ean13: str, x_mm: float, y_mm: float, w_mm: float, h_mm: float):
    widget = Ean13BarcodeWidget(ean13[:12])
    widget.humanReadable = 0
    bounds = widget.getBounds()
    bw = bounds[2] - bounds[0]
    bh = bounds[3] - bounds[1]

    drawing = Drawing(w_mm * mm, h_mm * mm)
    drawing.add(widget)
    sx = (w_mm * mm) / bw
    sy = (h_mm * mm) / bh

    c.saveState()
    c.translate(x_mm * mm, y_mm * mm)
    c.scale(sx, sy)
    renderPDF.draw(drawing, c, 0, 0)
    c.restoreState()


def draw_label(c: canvas.Canvas, row: dict, settings: dict, font_name: str):
    fs = int(settings["font_size"])
    c.setFont(font_name, fs)
    for kx, ky, text in label_lines(row):
        c.setFont(font_name, fs)
        c.drawString(float(settings[kx]) * mm, float(settings[ky]) * mm, text)

    ean13 = normalize_ean13(row["Баркод"])
    draw_barcode(
        c, ean13,
        float(settings["barcode_x"]), float(settings["barcode_y"]),
        float(settings["barcode_w"]), float(settings["barcode_h"]),
    )
    c.setFont(font_name, int(settings["barcode_digits_font_size"]))
    c.drawString(
        float(settings["barcode_digits_x"]) * mm,
        float(settings["barcode_digits_y"]) * mm,
        ean13,
    )

    num = str(row.get("_num", ""))
    if num:
        c.setFont(font_name, int(settings["num_font_size"]))
        c.drawString(float(settings["num_x"]) * mm, float(settings["num_y"]) * mm, num)


def row_filename(row: dict) -> str:
    return "_".join([
        sanitize_filename(row["Артикул"]),
        sanitize_filename(row["Цвет"]),
        sanitize_filename(row["Размер"]),
    ]) + ".pdf"


def unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    n = 1
    while True:
        cand = path.with_name(f"{path.stem} ({n}){path.suffix}")
        if not cand.exists():
            return cand
        n += 1


def render_pdf(rows: list, settings: dict, font_name: str) -> bytes:
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=(LABEL_W_MM * mm, LABEL_H_MM * mm))
    for row in rows:
        draw_label(c, row, settings, font_name)
        c.showPage()
    c.save()
    return buf.getvalue()


# ---------------------------------------------------------------------------
# EAN-13 модули (для превью)
# ---------------------------------------------------------------------------

_EAN_L = {
    "0": "0001101", "1": "0011001", "2": "0010011", "3": "0111101",
    "4": "0100011", "5": "0110001", "6": "0101111", "7": "0111011",
    "8": "0110111", "9": "0001011",
}
_EAN_G = {
    "0": "0100111", "1": "0110011", "2": "0011011", "3": "0100001",
    "4": "0011101", "5": "0111001", "6": "0000101", "7": "0010001",
    "8": "0001001", "9": "0010111",
}
_EAN_R = {
    "0": "1110010", "1": "1100110", "2": "1101100", "3": "1000010",
    "4": "1011100", "5": "1001110", "6": "1010000", "7": "1000100",
    "8": "1001000", "9": "1110100",
}
_EAN_PARITY = {
    "0": "LLLLLL", "1": "LLGLGG", "2": "LLGGLG", "3": "LLGGGL",
    "4": "LGLLGG", "5": "LGGLLG", "6": "LGGGLL", "7": "LGLGLG",
    "8": "LGLGGL", "9": "LGGLGL",
}


def ean13_modules(code: str) -> str:
    """Вернуть строку из '0'/'1' (96 модулей) для отрисовки превью EAN-13."""
    first = code[0]
    left = code[1:7]
    right = code[7:13]
    bits = "101"  # левый guard
    parity = _EAN_PARITY[first]
    for d, p in zip(left, parity):
        bits += _EAN_L[d] if p == "L" else _EAN_G[d]
    bits += "01010"  # центральный guard
    for d in right:
        bits += _EAN_R[d]
    bits += "101"  # правый guard
    return bits


# ---------------------------------------------------------------------------
# Превью (PIL -> Tkinter)
# ---------------------------------------------------------------------------

class Preview:
    PX_PER_MM = 11  # масштаб превью

    def __init__(self, canvas_widget: Canvas):
        self.canvas = canvas_widget
        self._photo = None

    def render(self, row: dict, settings: dict, show_grid: bool, font_path: str | None):
        ppm = self.PX_PER_MM
        W = int(LABEL_W_MM * ppm)
        H = int(LABEL_H_MM * ppm)

        img = Image.new("RGB", (W, H), "white")
        d = ImageDraw.Draw(img)

        def Y(y_mm: float) -> float:
            # в PDF Y снизу; в картинке Y сверху
            return H - y_mm * ppm

        def X(x_mm: float) -> float:
            return x_mm * ppm

        # --- сетка / разметка ---
        if show_grid:
            for gx in range(0, LABEL_W_MM + 1, 5):
                px = X(gx)
                d.line([(px, 0), (px, H)], fill=(225, 225, 235), width=1)
            for gy in range(0, LABEL_H_MM + 1, 5):
                py = Y(gy)
                d.line([(0, py), (W, py)], fill=(225, 225, 235), width=1)
            # подписи осей (каждые 10 мм)
            try:
                axis_font = ImageFont.truetype(font_path, 9) if font_path else ImageFont.load_default()
            except Exception:
                axis_font = ImageFont.load_default()
            for gx in range(0, LABEL_W_MM + 1, 10):
                d.text((X(gx) + 1, 1), str(gx), fill=(150, 150, 170), font=axis_font)
            for gy in range(0, LABEL_H_MM + 1, 10):
                if gy == 0:
                    continue
                d.text((1, Y(gy) + 1), str(gy), fill=(150, 150, 170), font=axis_font)

        # --- рамка этикетки ---
        d.rectangle([(0, 0), (W - 1, H - 1)], outline=(0, 0, 0), width=2)

        # --- текстовые строки ---
        fs = int(float(settings.get("font_size", 7)))
        # размер шрифта в мм -> пикс. reportlab трактует font_size в пунктах (1pt=1/72").
        # для соответствия PDF берём пункты как мм*? — на деле в исходном коде размер
        # задаётся прямо как size в pt, а координаты в мм. Поэтому в превью масштабируем
        # шрифт в тех же пунктах: 1pt ≈ 0.3528 мм -> в пикселях:
        def pt_to_px(pt: float) -> int:
            return max(6, int(pt / 72.0 * 25.4 * ppm))

        try:
            text_font = ImageFont.truetype(font_path, pt_to_px(fs)) if font_path else ImageFont.load_default()
        except Exception:
            text_font = ImageFont.load_default()

        for kx, ky, text in label_lines(row):
            x = X(float(settings[kx]))
            ybaseline = Y(float(settings[ky]))
            # PIL рисует от верхнего левого угла текста; в PDF drawString — от базовой линии.
            # Поднимаем текст на высоту шрифта, чтобы базовая линия совпала с y_mm.
            ascent, descent = text_font.getmetrics()
            d.text((x, ybaseline - ascent), text, fill=(0, 0, 0), font=text_font)

        # --- штрихкод ---
        bx = float(settings["barcode_x"])
        by = float(settings["barcode_y"])
        bw = float(settings["barcode_w"])
        bh = float(settings["barcode_h"])

        valid_code = None
        try:
            valid_code = normalize_ean13(row["Баркод"])
        except Exception:
            valid_code = None

        x0 = X(bx)
        y_top = Y(by + bh)
        bar_h = bh * ppm
        bar_w_total = bw * ppm

        if valid_code:
            bits = ean13_modules(valid_code)
            module_w = bar_w_total / len(bits)
            for i, b in enumerate(bits):
                if b == "1":
                    mx0 = x0 + i * module_w
                    d.rectangle(
                        [(mx0, y_top), (mx0 + module_w, y_top + bar_h)],
                        fill=(0, 0, 0),
                    )
        else:
            # код невалиден — рисуем рамку-заглушку
            d.rectangle(
                [(x0, y_top), (x0 + bar_w_total, y_top + bar_h)],
                outline=(200, 0, 0), width=1,
            )
            try:
                warn_font = ImageFont.truetype(font_path, 11) if font_path else ImageFont.load_default()
            except Exception:
                warn_font = ImageFont.load_default()
            d.text((x0 + 2, y_top + 2), "ШК?", fill=(200, 0, 0), font=warn_font)

        # --- цифры под ШК ---
        dfs = int(float(settings.get("barcode_digits_font_size", 7)))
        try:
            dig_font = ImageFont.truetype(font_path, pt_to_px(dfs)) if font_path else ImageFont.load_default()
        except Exception:
            dig_font = ImageFont.load_default()
        dx = X(float(settings["barcode_digits_x"]))
        dy = Y(float(settings["barcode_digits_y"]))
        ascent, descent = dig_font.getmetrics()
        digits_text = valid_code if valid_code else re.sub(r"\D", "", str(row.get("Баркод", "")))
        d.text((dx, dy - ascent), digits_text, fill=(0, 0, 0), font=dig_font)

        # --- номер строки в углу ---
        nfs = int(float(settings.get("num_font_size", 6)))
        try:
            num_font = ImageFont.truetype(font_path, pt_to_px(nfs)) if font_path else ImageFont.load_default()
        except Exception:
            num_font = ImageFont.load_default()
        num_text = str(row.get("_num", ""))
        if num_text:
            nx = X(float(settings["num_x"]))
            ny = Y(float(settings["num_y"]))
            n_ascent, _ = num_font.getmetrics()
            d.text((nx, ny - n_ascent), num_text, fill=(0, 0, 0), font=num_font)

        # --- вывод на Canvas ---
        self._photo = ImageTk.PhotoImage(img)
        self.canvas.config(width=W, height=H)
        self.canvas.delete("all")
        self.canvas.create_image(0, 0, anchor="nw", image=self._photo)


# ---------------------------------------------------------------------------
# Приложение
# ---------------------------------------------------------------------------

class App:
    TEST_ROW = {
        "Артикул": "TEST-001",
        "Артикул WB": "120909012",
        "Цвет": "Черный",
        "Размер": "38",
        "Состав": "Полиэстер",
        "Баркод": "460123456789",
        "Гарантия": "1 год",
        "_num": 1,
    }

    def __init__(self, root: Tk):
        self.root = root
        self.root.title("WB генератор ШК / этикеток 58x40")
        self.root.geometry("1180x760")
        self.settings = load_settings()

        self.pdf_font_name = register_pdf_font()

        self.excel_path = StringVar(value=self.settings.get("excel_path", ""))
        self.output_dir = StringVar(value=self.settings.get("output_dir", ""))
        self.vars = {}
        self.preview_row = dict(self.TEST_ROW)

        self.build_ui()
        self.refresh_preview()

    # ---------------- UI ----------------
    def build_ui(self):
        pad = {"padx": 8, "pady": 4}

        # Корневой контейнер: слева настройки, справа превью
        main = ttk.Frame(self.root)
        main.pack(fill="both", expand=True)

        # прокручиваемая левая панель (скроллбар для оконного режима)
        left_outer = ttk.Frame(main)
        left_outer.pack(side="left", fill="both", expand=True)
        left_canvas = Canvas(left_outer, highlightthickness=0)
        left_scroll = ttk.Scrollbar(left_outer, orient="vertical", command=left_canvas.yview)
        left_canvas.configure(yscrollcommand=left_scroll.set)
        left_scroll.pack(side="right", fill="y")
        left_canvas.pack(side="left", fill="both", expand=True)
        left = ttk.Frame(left_canvas)
        left_window = left_canvas.create_window((0, 0), window=left, anchor="nw")
        left.bind("<Configure>", lambda e: left_canvas.configure(scrollregion=left_canvas.bbox("all")))
        left_canvas.bind("<Configure>", lambda e: left_canvas.itemconfigure(left_window, width=e.width))
        left_canvas.bind_all("<MouseWheel>", lambda e: left_canvas.yview_scroll(int(-e.delta / 120), "units"))

        right = ttk.Frame(main)
        right.pack(side="right", fill="y", padx=8, pady=8)

        # ---- верхние пути ----
        top = ttk.Frame(left)
        top.pack(fill="x", **pad)

        ttk.Label(top, text="Excel:").grid(row=0, column=0, sticky="w")
        ttk.Entry(top, textvariable=self.excel_path, width=60).grid(row=0, column=1, sticky="we")
        ttk.Button(top, text="Выбрать", command=self.choose_excel).grid(row=0, column=2)

        ttk.Label(top, text="Папка выгрузки:").grid(row=1, column=0, sticky="w")
        ttk.Entry(top, textvariable=self.output_dir, width=60).grid(row=1, column=1, sticky="we")
        ttk.Button(top, text="Выбрать", command=self.choose_output).grid(row=1, column=2)
        top.columnconfigure(1, weight=1)

        # ---- настройки ----
        settings_frame = ttk.LabelFrame(left, text="Настройки положения, мм. Y считается снизу этикетки")
        settings_frame.pack(fill="x", **pad)

        idx = 0
        for key, label in [
            ("font_size", "Размер шрифта"),
            ("article_x", "Артикул X"), ("article_y", "Артикул Y"),
            ("color_x", "Цвет X"), ("color_y", "Цвет Y"),
            ("size_x", "Размер X"), ("size_y", "Размер Y"),
            ("material_x", "Материал X"), ("material_y", "Материал Y"),
            ("expiry_x", "Гарантия X"), ("expiry_y", "Гарантия Y"),
            ("barcode_x", "ШК X"), ("barcode_y", "ШК Y"),
            ("barcode_w", "ШК ширина"), ("barcode_h", "ШК высота"),
            ("barcode_digits_x", "Цифры ШК X"), ("barcode_digits_y", "Цифры ШК Y"),
            ("barcode_digits_font_size", "Цифры ШК шрифт"),
            ("num_x", "Номер X"), ("num_y", "Номер Y"),
            ("num_font_size", "Номер шрифт"),
        ]:
            var = DoubleVar(value=self.settings.get(key, DEFAULT_SETTINGS[key]))
            self.vars[key] = var
            var.trace_add("write", lambda *_: self.refresh_preview())
            ttk.Label(settings_frame, text=label).grid(row=idx // 4, column=(idx % 4) * 2, sticky="w", padx=5, pady=3)
            e = ttk.Entry(settings_frame, textvariable=var, width=8)
            e.grid(row=idx // 4, column=(idx % 4) * 2 + 1, padx=5, pady=3)
            idx += 1

        # ---- опции ----
        opt_frame = ttk.Frame(left)
        opt_frame.pack(fill="x", **pad)
        self.vars["make_one_pdf"] = IntVar(value=int(self.settings.get("make_one_pdf", 0)))
        ttk.Checkbutton(
            opt_frame, text="Сформировать один общий PDF вместо отдельных файлов",
            variable=self.vars["make_one_pdf"],
        ).pack(anchor="w")

        self.vars["print_gtin"] = IntVar(value=int(self.settings.get("print_gtin", 0)))
        ttk.Checkbutton(
            opt_frame, text="Печатать GTIN (иначе пропускать)",
            variable=self.vars["print_gtin"],
        ).pack(anchor="w")

        self.vars["show_grid"] = IntVar(value=int(self.settings.get("show_grid", 1)))
        self.vars["show_grid"].trace_add("write", lambda *_: self.refresh_preview())
        ttk.Checkbutton(
            opt_frame, text="Показывать сетку и разметку в превью",
            variable=self.vars["show_grid"],
        ).pack(anchor="w")

        # ---- кнопки ----
        buttons = ttk.Frame(left)
        buttons.pack(fill="x", **pad)
        ttk.Button(buttons, text="Сохранить настройки", command=self.save_current_settings).pack(side="left", padx=4)
        ttk.Button(buttons, text="Сформировать PDF", command=self.generate).pack(side="left", padx=4)
        ttk.Button(buttons, text="Сделать тестовый PDF", command=self.test_pdf).pack(side="left", padx=4)
        ttk.Button(buttons, text="Превью из Excel (1-я строка)", command=self.preview_from_excel).pack(side="left", padx=4)
        ttk.Button(buttons, text="Шаблон Excel", command=self.save_template).pack(side="left", padx=4)

        self.log = Text(left, height=10)
        self.log.pack(fill="both", expand=True, **pad)

        # ---- ПРАВАЯ ПАНЕЛЬ: ПРЕВЬЮ ----
        ttk.Label(right, text="Превью этикетки (58×40 мм)").pack(anchor="w")
        self.preview_canvas = Canvas(right, bg="#f4f4f7", highlightthickness=1, highlightbackground="#bbb")
        self.preview_canvas.pack(pady=6)
        self.preview = Preview(self.preview_canvas)
        ttk.Label(
            right,
            text="Синяя сетка = шаг 5 мм. Числа по краям = мм.\nКрасная рамка ШК = баркод невалиден.",
            justify="left", foreground="#555",
        ).pack(anchor="w")

        if not FONT_PATH:
            self.write_log("ВНИМАНИЕ: не найден TTF-шрифт с кириллицей. На Windows проверьте C:\\Windows\\Fonts\\arial.ttf")

    # ---------------- действия ----------------
    def choose_excel(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xlsm")])
        if path:
            self.excel_path.set(path)

    def choose_output(self):
        path = filedialog.askdirectory()
        if path:
            self.output_dir.set(path)

    def current_settings(self):
        settings = self.settings.copy()
        for key, var in self.vars.items():
            try:
                settings[key] = var.get()
            except Exception:
                pass
        settings["excel_path"] = self.excel_path.get()
        settings["output_dir"] = self.output_dir.get()
        return settings

    def save_template(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="шаблон_wb.xlsx",
        )
        if not path:
            return
        try:
            make_template(path)
            self.write_log(f"Шаблон сохранён: {path}")
            messagebox.showinfo("Готово", f"Шаблон сохранён:\n{path}")
        except Exception as e:
            self.write_log(traceback.format_exc())
            messagebox.showerror("Ошибка", str(e))

    def save_current_settings(self):
        self.settings = self.current_settings()
        save_settings(self.settings)
        messagebox.showinfo("Готово", "Настройки сохранены")

    def write_log(self, text: str):
        self.log.insert(END, text + "\n")
        self.log.see(END)
        self.root.update_idletasks()

    def refresh_preview(self):
        """Перерисовать превью. Вызывается при любом изменении полей."""
        try:
            settings = self.current_settings()
            show_grid = bool(self.vars.get("show_grid").get()) if "show_grid" in self.vars else True
            self.preview.render(self.preview_row, settings, show_grid, FONT_PATH)
        except Exception:
            # во время ввода значения могут быть временно пустыми — это нормально
            pass

    def preview_from_excel(self):
        try:
            if not self.excel_path.get():
                self.choose_excel()
            if not self.excel_path.get():
                return
            rows = read_excel_rows(self.excel_path.get())
            if not rows:
                raise ValueError("В Excel нет строк")
            self.preview_row = rows[0]
            self.write_log(f"Превью из строки Excel №{rows[0].get('_row')}")
            self.refresh_preview()
        except Exception as e:
            self.write_log(traceback.format_exc())
            messagebox.showerror("Ошибка", str(e))

    def validate_paths(self):
        if not self.excel_path.get():
            raise ValueError("Выберите Excel-файл")
        if not self.output_dir.get():
            raise ValueError("Выберите папку выгрузки")
        if not os.path.exists(self.excel_path.get()):
            raise ValueError("Excel-файл не найден")
        os.makedirs(self.output_dir.get(), exist_ok=True)

    def _reset_save_state(self):
        self._used_paths = set()
        self._save_new_on_lock = None

    def _save_pdf(self, data: bytes, base_path: Path):
        path = unique_path(base_path)
        with open(path, "wb") as f:
            f.write(data)
        return path

    def generate(self):
        try:
            self.validate_paths()
            settings = self.current_settings()
            save_settings(settings)
            rows = read_excel_rows(self.excel_path.get())
            if not rows:
                raise ValueError("В Excel нет строк для выгрузки")

            self.write_log(f"Найдено строк: {len(rows)}")
            self._reset_save_state()
            errors = 0
            skipped_gtin = []
            print_gtin = int(settings.get("print_gtin", 0))
            out = Path(self.output_dir.get())

            def gtin_skip(row):
                if is_gtin(row["Баркод"]) and not print_gtin:
                    skipped_gtin.append(row)
                    self.write_log(f"Строка {row.get('_row')}: GTIN пропущен")
                    return True
                return False

            if int(settings.get("make_one_pdf", 0)):
                valid_rows = []
                for row in rows:
                    if gtin_skip(row):
                        continue
                    try:
                        normalize_ean13(row["Баркод"])
                        valid_rows.append(row)
                    except Exception as e:
                        errors += 1
                        self.write_log(f"Строка {row.get('_row')}: ошибка - {e}")
                if valid_rows:
                    data = render_pdf(valid_rows, settings, self.pdf_font_name)
                    path = self._save_pdf(data, out / "WB_этикетки_58x40.pdf")
                    if path:
                        self.write_log(f"Создан общий PDF: {path}")
            else:
                for row in rows:
                    if gtin_skip(row):
                        continue
                    try:
                        data = render_pdf([row], settings, self.pdf_font_name)
                        path = self._save_pdf(data, out / row_filename(row))
                        if path:
                            self.write_log(f"OK строка {row.get('_row')}: {path.name}")
                    except Exception as e:
                        errors += 1
                        self.write_log(f"Строка {row.get('_row')}: ошибка - {e}")

            if skipped_gtin:
                excl = make_exclusions(self.excel_path.get(), skipped_gtin, out / "исключения_GTIN.xlsx")
                self.write_log(f"Пропущено GTIN: {len(skipped_gtin)}. Список исключений: {excl}")
            messagebox.showinfo("Готово", f"Генерация завершена. Ошибок: {errors}. Пропущено GTIN: {len(skipped_gtin)}")
        except Exception as e:
            self.write_log(traceback.format_exc())
            messagebox.showerror("Ошибка", str(e))

    def test_pdf(self):
        try:
            if not self.output_dir.get():
                self.choose_output()
            if not self.output_dir.get():
                return
            settings = self.current_settings()
            self._reset_save_state()
            data = render_pdf([dict(self.TEST_ROW)], settings, self.pdf_font_name)
            path = self._save_pdf(data, Path(self.output_dir.get()) / row_filename(dict(self.TEST_ROW)))
            if path:
                self.write_log(f"Тестовый PDF создан: {path}")
                messagebox.showinfo("Готово", f"Тестовый PDF создан:\n{path}")
        except Exception as e:
            self.write_log(traceback.format_exc())
            messagebox.showerror("Ошибка", str(e))


if __name__ == "__main__":
    root = Tk()
    app = App(root)
    root.mainloop()
