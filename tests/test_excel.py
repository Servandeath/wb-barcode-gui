import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

import wb_barcode_gui as app


class ExcelTests(unittest.TestCase):
    def test_missing_required_columns_are_reported(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "invalid.xlsx"
            workbook = Workbook()
            workbook.active.append(["Артикул", "Баркод"])
            workbook.save(path)
            with self.assertRaisesRegex(ValueError, "В Excel нет колонок"):
                app.read_excel_rows(str(path))

    def test_exclusions_highlight_skipped_source_rows(self):
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp) / "source.xlsx"
            app.make_template(str(source))
            output = app.make_exclusions(
                str(source), [{"_row": 2}], Path(tmp) / "excluded.xlsx"
            )
            sheet = load_workbook(output).active
            self.assertEqual(sheet["A2"].fill.fill_type, "solid")
            self.assertEqual(sheet["A2"].fill.fgColor.rgb, "00FFFF00")

    def test_empty_rows_are_ignored(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "rows.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(app.REQUIRED_COLUMNS)
            sheet.append([None] * len(app.REQUIRED_COLUMNS))
            workbook.save(path)
            self.assertEqual(app.read_excel_rows(str(path)), [])


if __name__ == "__main__":
    unittest.main()
